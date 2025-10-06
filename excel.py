import os
import argparse
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# Reuse DB helper & exports dir from analytics.py
from analytics import fetch_df, EXPORTS_DIR

# ---------------- Excel helpers ----------------
def _fit_column_widths(ws, df: pd.DataFrame, min_w: int = 8, max_w: int = 48):
    for j, col in enumerate(df.columns, start=1):
        header_len = len(str(col))
        sample_vals = df[col].astype(str).head(200).tolist()
        body_len = max((len(v) for v in sample_vals), default=0)
        width = max(header_len, body_len) + 2
        width = max(min_w, min(width, max_w))
        ws.column_dimensions[get_column_letter(j)].width = width

def _add_numeric_colorscale(ws, df: pd.DataFrame, start_row: int = 2):
    num_cols = [i for i, c in enumerate(df.columns, start=1)
                if pd.api.types.is_numeric_dtype(df[c])]
    if not num_cols or ws.max_row < start_row:
        return
    rule = ColorScaleRule(
        start_type="min", start_color="FFAA0000",
        mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
        end_type="max", end_color="FF00AA00"
    )
    for j in num_cols:
        letter = get_column_letter(j)
        rng = f"{letter}{start_row}:{letter}{ws.max_row}"
        ws.conditional_formatting.add(rng, rule)

def _percentify(ws, df: pd.DataFrame, columns, start_row: int = 2):
    idx = {c: i + 1 for i, c in enumerate(df.columns)}
    for col in columns:
        if col not in idx: continue
        letter = get_column_letter(idx[col])
        for r in range(start_row, ws.max_row + 1):
            ws[f"{letter}{r}"].number_format = "0.00%"

def _currency(ws, df: pd.DataFrame, columns, start_row: int = 2):
    idx = {c: i + 1 for i, c in enumerate(df.columns)}
    for col in columns:
        if col not in idx: continue
        letter = get_column_letter(idx[col])
        for r in range(start_row, ws.max_row + 1):
            ws[f"{letter}{r}"].number_format = "#,##0.00"

def _datefmt(ws, df: pd.DataFrame, columns, fmt="yyyy-mm-dd", start_row: int = 2):
    idx = {c: i + 1 for i, c in enumerate(df.columns)}
    for col in columns:
        if col not in idx: continue
        letter = get_column_letter(idx[col])
        for r in range(start_row, ws.max_row + 1):
            ws[f"{letter}{r}"].number_format = fmt

def _safe_sheet(name: str) -> str:
    bad = '[]:*?/\\'
    clean = ''.join('_' if ch in bad else ch for ch in name)[:31]
    return clean or "Sheet"

# ---------------- Profiling helpers ----------------
def _series_top_k(s: pd.Series, k: int = 3):
    vc = s.astype(str).value_counts(dropna=True)
    top = vc.head(k)
    total = len(s)
    rows = []
    for idx, cnt in top.items():
        share = (cnt / total) * 100 if total else 0.0
        rows.append((idx, int(cnt), round(share, 2)))
    while len(rows) < k:
        rows.append(("", 0, 0.0))
    return rows

def _profile_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=[
            "column", "dtype", "non_null", "missing", "missing_%", "unique",
            "min", "p25", "median", "p75", "max", "mean", "std",
            "top1", "top1_cnt", "top1_share%", "top2", "top2_cnt", "top2_share%",
            "top3", "top3_cnt", "top3_share%"
        ])
    rows = []
    for col in df.columns:
        s = df[col]
        dtype = str(s.dtype)
        non_null = int(s.notna().sum())
        missing = int(s.isna().sum())
        missing_pct = (missing / len(s) * 100) if len(s) else 0.0
        nunique = int(s.nunique(dropna=True))
        stats = {
            "min": None, "p25": None, "median": None, "p75": None, "max": None,
            "mean": None, "std": None,
            "top1": None, "top1_cnt": None, "top1_share%": None,
            "top2": None, "top2_cnt": None, "top2_share%": None,
            "top3": None, "top3_cnt": None, "top3_share%": None,
        }
        if pd.api.types.is_numeric_dtype(s):
            sn = pd.to_numeric(s, errors="coerce")
            if sn.notna().any():
                q = sn.quantile([0.25, 0.5, 0.75])
                stats.update({
                    "min": float(sn.min()),
                    "p25": float(q.get(0.25, np.nan)),
                    "median": float(q.get(0.5, np.nan)),
                    "p75": float(q.get(0.75, np.nan)),
                    "max": float(sn.max()),
                    "mean": float(sn.mean()),
                    "std": float(sn.std(ddof=1)) if sn.count() > 1 else 0.0,
                })
        else:
            (t1, c1, p1), (t2, c2, p2), (t3, c3, p3) = _series_top_k(s, 3)
            stats.update({
                "top1": t1, "top1_cnt": c1, "top1_share%": p1,
                "top2": t2, "top2_cnt": c2, "top2_share%": p2,
                "top3": t3, "top3_cnt": c3, "top3_share%": p3,
            })
        rows.append({
            "column": col, "dtype": dtype, "non_null": non_null,
            "missing": missing, "missing_%": round(missing_pct, 2),
            "unique": nunique, **stats
        })
    front = ["column", "dtype", "non_null", "missing", "missing_%", "unique"]
    num_stats = ["min", "p25", "median", "p75", "max", "mean", "std"]
    top_stats = ["top1", "top1_cnt", "top1_share%", "top2", "top2_cnt", "top2_share%",
                 "top3", "top3_cnt", "top3_share%"]
    prof = pd.DataFrame(rows)
    cols = [*front, *num_stats, *top_stats]
    return prof[[c for c in cols if c in prof.columns]]

# ---------------- Dataset builders + summaries ----------------
def build_pie():
    df = fetch_df("pie_revenue_by_category")
    df["revenue"] = pd.to_numeric(df.get("revenue", 0), errors="coerce").fillna(0.0)
    df = df.sort_values("revenue", ascending=False).reset_index(drop=True)
    total = df["revenue"].sum()
    df["share"] = (df["revenue"] / total) if total else 0.0
    df["rank"]  = df["revenue"].rank(method="first", ascending=False).astype(int)
    df["cum_share"] = df["share"].cumsum()
    df = df[["rank", "category", "revenue", "share", "cum_share"]]
    return df

def summary_pie(df):
    total_revenue = float(df["revenue"].sum()) if len(df) else 0.0
    n_categories = int(df["category"].nunique()) if len(df) else 0
    if len(df):
        top = df.iloc[0]
        top_cat, top_rev, top_share = str(top["category"]), float(top["revenue"]), float(top["share"])
        p = df["share"].to_numpy(dtype=float); p = p[(p > 0) & np.isfinite(p)]
        hhi = float(np.sum(p**2)) if p.size else 0.0
        shannon = float(-(p * np.log2(p)).sum()) if p.size else 0.0
    else:
        top_cat = ""; top_rev = top_share = hhi = shannon = 0.0
    rows = [
        {"Metric":"Total revenue (top slice)","Value":total_revenue,"Note":"Sum of rows (top-10 categories)."},
        {"Metric":"Categories in table","Value":n_categories,"Note":"Row count in this slice."},
        {"Metric":"Top-1 category","Value":top_cat,"Note":"Max revenue category."},
        {"Metric":"Top-1 revenue","Value":top_rev,"Note":"Revenue of top category."},
        {"Metric":"Top-1 share","Value":top_share,"Note":"Share of top category within this table."},
        {"Metric":"HHI (0–1)","Value":hhi,"Note":"Higher → more concentrated."},
        {"Metric":"Shannon entropy (bits)","Value":shannon,"Note":"Higher → more even."},
    ]
    return pd.DataFrame(rows), {"percent_cols_in_value_col_by_label":["share"]}

def build_top_sellers():
    df = fetch_df("bar_top_sellers_by_revenue")
    df.rename(columns={"seller_id":"seller"}, inplace=True)
    df["revenue"] = pd.to_numeric(df["revenue"], errors="coerce").fillna(0.0)
    df = df.sort_values("revenue", ascending=False).reset_index(drop=True)
    total = df["revenue"].sum()
    df["share"] = (df["revenue"]/total) if total else 0.0
    df["rank"] = df["revenue"].rank(method="first", ascending=False).astype(int)
    df = df[["rank", "seller", "revenue", "share"]]
    return df

def summary_top_sellers(df):
    total = float(df["revenue"].sum()) if len(df) else 0.0
    n = int(df["seller"].nunique()) if len(df) else 0
    top_row = df.iloc[0] if len(df) else None
    rows = [
        {"Metric":"Total revenue (top sellers)","Value":total,"Note":"Sum of this table (top-10)."},
        {"Metric":"Count of sellers","Value":n,"Note":"Unique seller names in the slice."},
        {"Metric":"Top seller","Value":(top_row["seller"] if top_row is not None else ""), "Note":"Highest revenue seller."},
        {"Metric":"Top seller revenue","Value":(float(top_row["revenue"]) if top_row is not None else 0.0), "Note":"Absolute revenue."},
        {"Metric":"Top seller share","Value":(float(top_row["share"]) if top_row is not None else 0.0), "Note":"Share within this slice."},
    ]
    return pd.DataFrame(rows), {"percent_cols_in_value_col_by_label":["share"]}

def build_barh_review():
    df = fetch_df("barh_avg_review_by_category")
    df.rename(columns={"avg_score":"avg_score","n_reviews":"n_reviews"}, inplace=True)
    df["avg_score"] = pd.to_numeric(df["avg_score"], errors="coerce")
    df["n_reviews"] = pd.to_numeric(df["n_reviews"], errors="coerce").astype("Int64")
    df = df.sort_values(["avg_score","n_reviews"], ascending=[False, False]).reset_index(drop=True)
    df["rank"] = np.arange(1, len(df)+1, dtype=int)
    df = df[["rank","category","avg_score","n_reviews"]]
    return df

def summary_barh_review(df):
    if len(df):
        top = df.iloc[0]
        rows = [
            {"Metric":"Top category by avg_score","Value":str(top["category"]), "Note":"Highest avg_score (proxy)."},
            {"Metric":"Top avg_score","Value":float(top["avg_score"]), "Note":"Proxy score (from unit price)."},
            {"Metric":"Reviews (top row)","Value":int(top["n_reviews"]), "Note":"Row’s n_reviews."},
            {"Metric":"Rows in table","Value":int(len(df)), "Note":"Categories retained (≥50 reviews)."},
        ]
    else:
        rows = [{"Metric":"Rows in table","Value":0,"Note":"No data"}]
    return pd.DataFrame(rows), {}

def build_daily_rev():
    df = fetch_df("line_daily_revenue_2010_2014", parse_dates=["day"])
    df["revenue"] = pd.to_numeric(df["revenue"], errors="coerce").fillna(0.0)
    s = df["revenue"]
    df["ma7"] = s.rolling(7, min_periods=1).mean()
    df["p25"] = s.rolling(7, min_periods=1).quantile(0.25)
    df["p75"] = s.rolling(7, min_periods=1).quantile(0.75)
    df = df[["day","revenue","ma7","p25","p75"]]
    return df

def summary_daily_rev(df):
    if len(df):
        total = float(df["revenue"].sum())
        avg_day = float(df["revenue"].mean())
        peak_idx = int(df["revenue"].idxmax())
        peak_day = df.loc[peak_idx, "day"]; peak_val = float(df.loc[peak_idx, "revenue"])
        rows = [
            {"Metric":"Date range","Value":f"{df['day'].min().date()} → {df['day'].max().date()}","Note":"From SQL window"},
            {"Metric":"Total revenue","Value":total,"Note":"Sum over period"},
            {"Metric":"Average daily revenue","Value":avg_day,"Note":"Mean(revenue)"},
            {"Metric":"Peak day","Value":str(peak_day.date()),"Note":"Max(revenue) day"},
            {"Metric":"Peak revenue","Value":peak_val,"Note":"Max(revenue)"},
            {"Metric":"Std dev (daily)","Value":float(df['revenue'].std(ddof=1)), "Note":"Volatility"},
        ]
    else:
        rows = [{"Metric":"Rows","Value":0,"Note":"No data"}]
    return pd.DataFrame(rows), {"date_cols_in_value_col_by_label":["Date range"]}

def build_hist_orders():
    df = fetch_df("hist_order_value")
    df["order_value"] = pd.to_numeric(df["order_value"], errors="coerce").dropna()
    df = df[["order_value"]].reset_index(drop=True)
    return df

def _gini(x: np.ndarray) -> float:
    if x.size == 0: return 0.0
    x = x.astype(float)
    if np.all(x == 0): return 0.0
    x = np.sort(x); n = x.size; cumx = np.cumsum(x)
    g = (n + 1 - 2 * (cumx.sum() / cumx[-1])) / n
    return float(g)

def summary_hist_orders(df):
    x = df["order_value"].to_numpy(dtype=float) if len(df) else np.array([])
    quant = (lambda q: float(np.quantile(x, q))) if x.size else (lambda q: 0.0)
    rows = [
        {"Metric":"Invoices","Value":int(len(x)), "Note":"Number of orders"},
        {"Metric":"Mean","Value":float(x.mean()) if x.size else 0.0, "Note":"Average order value"},
        {"Metric":"Median","Value":quant(0.5), "Note":"50th percentile"},
        {"Metric":"Std dev","Value":float(x.std(ddof=1)) if x.size>1 else 0.0, "Note":"Spread"},
        {"Metric":"Min","Value":float(x.min()) if x.size else 0.0, "Note":""},
        {"Metric":"P25","Value":quant(0.25), "Note":""},
        {"Metric":"P75","Value":quant(0.75), "Note":""},
        {"Metric":"P90","Value":quant(0.90), "Note":""},
        {"Metric":"P95","Value":quant(0.95), "Note":""},
        {"Metric":"Max","Value":float(x.max()) if x.size else 0.0, "Note":""},
        {"Metric":"Gini (0–1)","Value":_gini(x), "Note":"Inequality of order values"},
    ]
    return pd.DataFrame(rows), {}

def build_duration_genre():
    df = fetch_df("duration_by_genre_minutes")
    df["duration_min"] = pd.to_numeric(df["duration_min"], errors="coerce")
    df = df.dropna(subset=["duration_min"]).reset_index(drop=True)
    df = df[["genre","duration_min"]]
    return df

def summary_duration_genre(df):
    rows = []
    if len(df):
        rows.append({"Metric":"Rows","Value":int(len(df)),"Note":"Tracks with duration"})
        rows.append({"Metric":"Genres","Value":int(df['genre'].nunique()),"Note":"Unique genres"})
        rows.append({"Metric":"Mean duration (min)","Value":float(df['duration_min'].mean()),"Note":"Across all rows"})
        rows.append({"Metric":"Median duration (min)","Value":float(df['duration_min'].median()),"Note":"50th percentile"})
        rows.append({"Metric":"P95 duration (min)","Value":float(df['duration_min'].quantile(0.95)),"Note":"Long-tail"})
        # top genre by median duration
        med = df.groupby("genre")["duration_min"].median().sort_values(ascending=False)
        if len(med):
            rows.append({"Metric":"Top genre by median duration","Value":str(med.index[0]),"Note":"Genre-level median"})
            rows.append({"Metric":"That median (min)","Value":float(med.iloc[0]),"Note":""})
    else:
        rows.append({"Metric":"Rows","Value":0,"Note":"No data"})
    return pd.DataFrame(rows), {}

def build_timeslider():
    df = fetch_df("timeslider_monthly_revenue_by_country")
    # month is 'YYYY-MM' string; keep as str but also provide an ordering helper
    df["revenue"] = pd.to_numeric(df["revenue"], errors="coerce").fillna(0.0)
    # rank within month (highest first)
    df["rank_in_month"] = df.groupby("month")["revenue"].rank(method="first", ascending=False).astype(int)
    df = df.sort_values(["month","rank_in_month"]).reset_index(drop=True)
    return df[["month","country","revenue","rank_in_month"]]

def summary_timeslider(df):
    rows = []
    if len(df):
        rows.append({"Metric":"Months","Value":int(df['month'].nunique()),"Note":"Distinct months"})
        rows.append({"Metric":"Countries (in data)","Value":int(df['country'].nunique()),"Note":"Distinct country names"})
        rows.append({"Metric":"Total revenue","Value":float(df['revenue'].sum()),"Note":"Sum across all month-country rows"})
        # top month by total revenue
        msum = df.groupby("month")["revenue"].sum().sort_values(ascending=False)
        top_month, top_val = (msum.index[0], float(msum.iloc[0])) if len(msum) else ("", 0.0)
        rows.append({"Metric":"Top month","Value":top_month,"Note":"By total revenue"})
        rows.append({"Metric":"Top month revenue","Value":top_val,"Note":""})
    else:
        rows.append({"Metric":"Rows","Value":0,"Note":"No data"})
    return pd.DataFrame(rows), {}
# ---------------- Registry of datasets ----------------
DATASETS = {
    "pie_revenue_by_category": {
        "builder": build_pie, "summary": summary_pie,
        "sheet": "01_pie_revenue_by_category",
        "fmt_currency": ["revenue"], "fmt_percent": ["share","cum_share"],
        "fmt_dates": []
    },
    "bar_top_sellers_by_revenue": {
        "builder": build_top_sellers, "summary": summary_top_sellers,
        "sheet": "02_bar_top_sellers_by_revenue",
        "fmt_currency": ["revenue"], "fmt_percent": ["share"], "fmt_dates": []
    },
    "barh_avg_review_by_category": {
        "builder": build_barh_review, "summary": summary_barh_review,
        "sheet": "03_barh_avg_review_by_category",
        "fmt_currency": [], "fmt_percent": [], "fmt_dates": []
    },
    "line_daily_revenue_2010_2014": {
        "builder": build_daily_rev, "summary": summary_daily_rev,
        "sheet": "04_line_daily_revenue_2010_2014",
        "fmt_currency": ["revenue","ma7","p25","p75"],
        "fmt_percent": [], "fmt_dates": ["day"]
    },
    "hist_order_value": {
        "builder": build_hist_orders, "summary": summary_hist_orders,
        "sheet": "05_hist_order_value",
        "fmt_currency": ["order_value"], "fmt_percent": [], "fmt_dates": []
    },
    "duration_by_genre_minutes": {
        "builder": build_duration_genre, "summary": summary_duration_genre,
        "sheet": "06_duration_by_genre_minutes",
        "fmt_currency": [], "fmt_percent": [], "fmt_dates": []
    },
    "timeslider_monthly_revenue_by_country": {
        "builder": build_timeslider, "summary": summary_timeslider,
        "sheet": "07_timeslider_revenue_by_country",
        "fmt_currency": ["revenue"], "fmt_percent": [], "fmt_dates": []
    },
}

# ---------------- Export core ----------------
def _write_dataframe_sheet(writer, sheet_name: str, df: pd.DataFrame,
                           fmt_currency=None, fmt_percent=None, fmt_dates=None):
    sname = _safe_sheet(sheet_name)
    df.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    _fit_column_widths(ws, df); _add_numeric_colorscale(ws, df, start_row=2)
    if fmt_currency: _currency(ws, df, fmt_currency)
    if fmt_percent:  _percentify(ws, df, fmt_percent)
    if fmt_dates:    _datefmt(ws, df, fmt_dates)
    return sname

def _write_summary_sheet(writer, base: str, summary_df: pd.DataFrame,
                         percent_labels_in_value=None, date_labels_in_value=None):
    name = _safe_sheet(f"Summary - {base}")
    summary_df.to_excel(writer, sheet_name=name, index=False)
    ws = writer.sheets[name]
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    _fit_column_widths(ws, summary_df); _add_numeric_colorscale(ws, summary_df, start_row=2)
    # smart-format the 'Value' column based on label keywords
    if "Value" in summary_df.columns:
        val_idx = list(summary_df.columns).index("Value") + 1
        val_letter = get_column_letter(val_idx)
        for r in range(2, ws.max_row + 1):
            label = ws[f"A{r}"].value or ""
            cell  = ws[f"{val_letter}{r}"]
            lab = str(label).lower()
            if percent_labels_in_value and any(k in lab for k in percent_labels_in_value):
                cell.number_format = "0.00%"
            elif date_labels_in_value and any(k in lab for k in date_labels_in_value):
                cell.number_format = "yyyy-mm-dd"
            else:
                cell.number_format = "#,##0.00"
    return name

def _write_profile_sheet(writer, base: str, df: pd.DataFrame):
    name = _safe_sheet(f"Profile - {base}")
    prof = _profile_dataframe(df)
    prof.to_excel(writer, sheet_name=name, index=False)
    ws = writer.sheets[name]
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    _fit_column_widths(ws, prof); _add_numeric_colorscale(ws, prof, start_row=2)
    return name

def export_datasets(out_path: str, dataset_keys: list[str]):
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        index_rows = []
        for key in dataset_keys:
            spec = DATASETS[key]
            df = spec["builder"]()
            base = spec["sheet"]
            # Data sheet
            data_sheet = _write_dataframe_sheet(
                writer, base, df,
                fmt_currency=spec.get("fmt_currency"),
                fmt_percent=spec.get("fmt_percent"),
                fmt_dates=spec.get("fmt_dates")
            )
            # Summary
            summ_df, fmt_hints = spec["summary"](df)
            percent_labels = fmt_hints.get("percent_cols_in_value_col_by_label", []) if fmt_hints else []
            date_labels    = fmt_hints.get("date_cols_in_value_col_by_label", []) if fmt_hints else []
            summary_sheet = _write_summary_sheet(writer, base, summ_df,
                                                 percent_labels_in_value=percent_labels,
                                                 date_labels_in_value=date_labels)
            # Profile
            profile_sheet = _write_profile_sheet(writer, base, df)

            index_rows.append({
                "Dataset key": key,
                "Data sheet": data_sheet,
                "Summary sheet": summary_sheet,
                "Profile sheet": profile_sheet,
                "Rows": len(df),
                "Columns": len(df.columns),
            })

        # Master index
        idx = pd.DataFrame(index_rows)
        idx.to_excel(writer, sheet_name="Index", index=False)
        ws = writer.sheets["Index"]
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        _fit_column_widths(ws, idx); _add_numeric_colorscale(ws, idx, start_row=2)
    print(f"[EXPORT] Excel → {out_path}")
    return out_path

# ---------------- CLI ----------------
def main():
    ap = argparse.ArgumentParser(
        description="Export all analytics datasets (data-only) with rich summaries + profiles."
    )
    ap.add_argument("--out", default=os.path.join(EXPORTS_DIR, "analytics_data.xlsx"),
                    help="Output .xlsx path")
    ap.add_argument("--only", choices=list(DATASETS.keys()),
                    help="Export only a single dataset (key from registry).")
    ap.add_argument("--no_timeslider", action="store_true",
                    help="Exclude timeslider dataset from 'all'.")
    args = ap.parse_args()

    if args.only:
        keys = [args.only]
        out = export_datasets(args.out, keys)
    else:
        keys = list(DATASETS.keys())
        if args.no_timeslider and "timeslider_monthly_revenue_by_country" in keys:
            keys.remove("timeslider_monthly_revenue_by_country")
        out = export_datasets(args.out, keys)
    print(f"[REPORT] Excel → {out}")

if __name__ == "__main__":
    main()
